from openpyxl import Workbook, load_workbook
from pathlib import Path
import os


def ensure_userdata_dir():
    os.makedirs("userdata", exist_ok=True)


def get_user_file(callsign: str) -> Path:
    ensure_userdata_dir()
    return Path("userdata") / f"{callsign.upper()}.xlsx"


def save_to_excel(rows, path: Path):
    """rows: list of lists [country, callsign, qso_date, qsl_status]"""
    wb = Workbook()
    ws = wb.active
    ws.append(["Country", "Callsign", "QSO Date", "QSL Status"])

    for r in rows:
        ws.append(r)

    wb.save(path)
    return Path(path).resolve()


def load_from_excel(path: Path):
    p = Path(path)
    if not p.exists():
        return []

    wb = load_workbook(p)
    ws = wb.active

    loaded = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            loaded.append(list(row))

    return loaded
